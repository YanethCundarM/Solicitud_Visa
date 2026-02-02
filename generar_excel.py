#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de Formulario DS-160 en Excel
Para colombianos solicitando visa americana
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def crear_formulario_visa():
    wb = Workbook()
    wb.remove(wb.active)
    
    # Estilos
    header_fill = PatternFill(start_color="0052CC", end_color="0052CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    section_fill = PatternFill(start_color="E8F0FF", end_color="E8F0FF", fill_type="solid")
    section_font = Font(bold=True, color="0052CC", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ======================== HOJA 1: INSTRUCCIONES ========================
    ws_instrucciones = wb.create_sheet("Instrucciones", 0)
    ws_instrucciones.column_dimensions['A'].width = 100
    
    row = 1
    ws_instrucciones[f'A{row}'] = "FORMULARIO DE SOLICITUD DE VISA AMERICANA (DS-160)"
    ws_instrucciones[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_instrucciones[f'A{row}'].fill = header_fill
    ws_instrucciones.row_dimensions[row].height = 25
    row += 2
    
    instrucciones = [
        "INSTRUCCIONES GENERALES:",
        "",
        "1. IMPORTANTE: Este formulario es una herramienta de AYUDA solamente.",
        "   Debes completar el formulario oficial DS-160 en: https://ceac.state.gov/GenNIV/General/complete/complete_personal.aspx",
        "",
        "2. SEGURIDAD DE DATOS: Todos los datos se guardan LOCALMENTE en tu computadora.",
        "   No se envía información a ningún servidor externo.",
        "",
        "3. PRIVACIDAD: Protege este archivo. Contiene información personal sensible.",
        "",
        "4. COMPLETITUD: Completa TODAS las secciones marcadas con asterisco (*).",
        "",
        "5. PRECISIÓN: Verifica que toda la información coincida con tus documentos oficiales.",
        "",
        "6. HONESTIDAD: Proporcionar información falsa es FRAUDE y puede resultar en:",
        "   - Denegación de visa",
        "   - Prohibición de entrada a EE.UU.",
        "   - Procesos legales",
        "   - Antecedentes penales",
        "",
        "7. DOCUMENTOS REQUERIDOS:",
        "   □ Pasaporte válido (original, vigencia +6 meses)",
        "   □ Cédula de ciudadanía",
        "   □ Foto tamaño 5x5 cm (fondo blanco)",
        "   □ Comprobante de pago de visa (DS-160 fee)",
        "   □ Comprobantes económicos (extractos bancarios, cartas laborales)",
        "   □ Comprobante de vivienda en Colombia",
        "   □ Si es B1: Carta del empleador",
        "",
        "8. PROCESO:",
        "   a) Completa este formulario",
        "   b) Llena el DS-160 oficial",
        "   c) Paga la tarifa de solicitud ($160 USD aproximadamente)",
        "   d) Programa tu cita en la embajada (https://ais.usvisa-info.com/es-co/)",
        "   e) Lleva todos los documentos a tu entrevista",
        "",
        "9. CONTACTOS ÚTILES:",
        "   - Embajada de EE.UU. en Colombia: https://co.usembassy.gov/",
        "   - Visa Information: https://travel.state.gov/",
        "   - CEAC DS-160: https://ceac.state.gov/GenNIV/",
        "   - Programa de Citas: https://ais.usvisa-info.com/es-co/",
        "",
        "10. VIGENCIA:",
        "    - La visa B1/B2 típicamente es válida por 10 años (ajustable)",
        "    - Puedes permanecer en EE.UU. máximo 6 meses por entrada (para turismo)",
        "    - El funcionario en migración determina el tiempo de permanencia",
    ]
    
    for instruccion in instrucciones:
        ws_instrucciones[f'A{row}'] = instruccion
        if instruccion.startswith(tuple('0123456789')):
            ws_instrucciones[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
    
    # ======================== HOJA 2: INFORMACIÓN PERSONAL ========================
    ws_personal = wb.create_sheet("Información Personal", 1)
    ws_personal.column_dimensions['A'].width = 25
    ws_personal.column_dimensions['B'].width = 40
    ws_personal.column_dimensions['C'].width = 25
    ws_personal.column_dimensions['D'].width = 40
    
    row = 1
    ws_personal[f'A{row}'] = "SECCIÓN 1: INFORMACIÓN PERSONAL"
    ws_personal.merge_cells(f'A{row}:D{row}')
    ws_personal[f'A{row}'].font = header_font
    ws_personal[f'A{row}'].fill = header_fill
    ws_personal.row_dimensions[row].height = 20
    row += 2
    
    # Datos básicos
    campos = [
        ("Apellido *", ""),
        ("Nombre *", ""),
        ("Segundo Nombre", ""),
        ("Otros Apellidos", ""),
        ("Fecha de Nacimiento (DD/MM/YYYY) *", ""),
        ("Lugar de Nacimiento *", ""),
        ("Género (M/F) *", ""),
        ("Nacionalidad *", "Colombiano/a"),
        ("Cédula de Ciudadanía *", ""),
        ("", ""),
        ("INFORMACIÓN DE CONTACTO", ""),
        ("Correo Electrónico *", ""),
        ("Teléfono Principal (+57) *", ""),
        ("Teléfono Secundario", ""),
    ]
    
    for label, default in campos:
        if label and not label.startswith("INFORMACIÓN"):
            ws_personal[f'A{row}'] = label
            ws_personal[f'A{row}'].font = Font(bold=True)
            ws_personal[f'B{row}'] = default
            ws_personal[f'B{row}'].fill = PatternFill(start_color="FFFFFFE0", end_color="FFFFFFE0", fill_type="solid")
        elif label.startswith("INFORMACIÓN"):
            ws_personal[f'A{row}'] = label
            ws_personal.merge_cells(f'A{row}:D{row}')
            ws_personal[f'A{row}'].font = section_font
            ws_personal[f'A{row}'].fill = section_fill
        row += 1
    
    # ======================== HOJA 3: PASAPORTE ========================
    ws_pasaporte = wb.create_sheet("Pasaporte", 2)
    ws_pasaporte.column_dimensions['A'].width = 25
    ws_pasaporte.column_dimensions['B'].width = 40
    
    row = 1
    ws_pasaporte[f'A{row}'] = "SECCIÓN 2: INFORMACIÓN DE PASAPORTE"
    ws_pasaporte.merge_cells(f'A{row}:B{row}')
    ws_pasaporte[f'A{row}'].font = header_font
    ws_pasaporte[f'A{row}'].fill = header_fill
    ws_pasaporte.row_dimensions[row].height = 20
    row += 2
    
    campos_pasaporte = [
        ("Número de Pasaporte (sin espacios) *", ""),
        ("País de Emisión *", "Colombia"),
        ("Fecha de Emisión (DD/MM/YYYY) *", ""),
        ("Fecha de Vencimiento (DD/MM/YYYY) *", ""),
        ("Nota", "Debe ser válido por 6 meses más desde la fecha de solicitud"),
        ("", ""),
        ("OBSERVACIONES:", ""),
        ("Motivo de emisión anterior", ""),
        ("Pasaporte anterior (número)", ""),
    ]
    
    for label, default in campos_pasaporte:
        if label and not label.startswith("OBSERVACIONES"):
            ws_pasaporte[f'A{row}'] = label
            ws_pasaporte[f'A{row}'].font = Font(bold=True)
            ws_pasaporte[f'B{row}'] = default
            ws_pasaporte[f'B{row}'].fill = PatternFill(start_color="FFFFFFE0", end_color="FFFFFFE0", fill_type="solid")
        elif label.startswith("OBSERVACIONES"):
            ws_pasaporte[f'A{row}'] = label
            ws_pasaporte.merge_cells(f'A{row}:B{row}')
            ws_pasaporte[f'A{row}'].font = section_font
            ws_pasaporte[f'A{row}'].fill = section_fill
        row += 1
    
    # ======================== HOJA 4: ANTECEDENTES ========================
    ws_antecedentes = wb.create_sheet("Antecedentes", 3)
    ws_antecedentes.column_dimensions['A'].width = 40
    ws_antecedentes.column_dimensions['B'].width = 30
    
    row = 1
    ws_antecedentes[f'A{row}'] = "SECCIÓN 3: ANTECEDENTES Y VIAJES PREVIOS"
    ws_antecedentes.merge_cells(f'A{row}:B{row}')
    ws_antecedentes[f'A{row}'].font = header_font
    ws_antecedentes[f'A{row}'].fill = header_fill
    ws_antecedentes.row_dimensions[row].height = 20
    row += 2
    
    campos_antecedentes = [
        ("¿Ha viajado a EE.UU. antes? (Sí/No)", ""),
        ("Año de último viaje", ""),
        ("Propósito del viaje anterior", ""),
        ("", ""),
        ("¿Tiene antecedentes penales? (Sí/No)", "No"),
        ("Descripción de antecedentes (si aplica)", ""),
        ("", ""),
        ("¿Tiene visas anteriores de EE.UU.? (Sí/No)", ""),
        ("Tipo de visa anterior", ""),
        ("Años de validez", ""),
    ]
    
    for label, default in campos_antecedentes:
        if label:
            ws_antecedentes[f'A{row}'] = label
            ws_antecedentes[f'A{row}'].font = Font(bold=True)
            ws_antecedentes[f'B{row}'] = default
            ws_antecedentes[f'B{row}'].fill = PatternFill(start_color="FFFFFFE0", end_color="FFFFFFE0", fill_type="solid")
        row += 1
    
    # ======================== HOJA 5: VIAJE ========================
    ws_viaje = wb.create_sheet("Viaje", 4)
    ws_viaje.column_dimensions['A'].width = 25
    ws_viaje.column_dimensions['B'].width = 40
    ws_viaje.column_dimensions['C'].width = 25
    ws_viaje.column_dimensions['D'].width = 40
    
    row = 1
    ws_viaje[f'A{row}'] = "SECCIÓN 4: INFORMACIÓN DEL VIAJE"
    ws_viaje.merge_cells(f'A{row}:D{row}')
    ws_viaje[f'A{row}'].font = header_font
    ws_viaje[f'A{row}'].fill = header_fill
    ws_viaje.row_dimensions[row].height = 20
    row += 2
    
    campos_viaje = [
        ("Tipo de Visa (B1, B2, B1/B2, F1, etc) *", ""),
        ("Fecha de Llegada Planeada (DD/MM/YYYY) *", ""),
        ("Ciudad/Estado de Destino *", ""),
        ("Duración del Viaje (semanas/meses) *", ""),
        ("Propósito Principal del Viaje *", ""),
        ("", ""),
        ("CONTACTO EN EE.UU.", ""),
        ("Empresa/Institución en EE.UU. *", ""),
        ("Dirección Completa *", ""),
        ("Ciudad y Estado *", ""),
        ("Teléfono", ""),
        ("Correo Electrónico", ""),
    ]
    
    for label, default in campos_viaje:
        if label and not label.startswith("CONTACTO"):
            ws_viaje[f'A{row}'] = label
            ws_viaje[f'A{row}'].font = Font(bold=True)
            ws_viaje[f'B{row}'] = default
            ws_viaje[f'B{row}'].fill = PatternFill(start_color="FFFFFFE0", end_color="FFFFFFE0", fill_type="solid")
        elif label.startswith("CONTACTO"):
            ws_viaje[f'A{row}'] = label
            ws_viaje.merge_cells(f'A{row}:D{row}')
            ws_viaje[f'A{row}'].font = section_font
            ws_viaje[f'A{row}'].fill = section_fill
        row += 1
    
    # ======================== HOJA 6: INFORMACIÓN LABORAL ========================
    ws_laboral = wb.create_sheet("Laboral", 5)
    ws_laboral.column_dimensions['A'].width = 30
    ws_laboral.column_dimensions['B'].width = 40
    
    row = 1
    ws_laboral[f'A{row}'] = "SECCIÓN 5: INFORMACIÓN LABORAL Y ECONÓMICA"
    ws_laboral.merge_cells(f'A{row}:B{row}')
    ws_laboral[f'A{row}'].font = header_font
    ws_laboral[f'A{row}'].fill = header_fill
    ws_laboral.row_dimensions[row].height = 20
    row += 2
    
    campos_laboral = [
        ("Estatus de Empleo *", ""),
        ("Ocupación/Cargo *", ""),
        ("Nombre del Empleador *", ""),
        ("Dirección del Empleador", ""),
        ("Años en el Empleo Actual", ""),
        ("", ""),
        ("INFORMACIÓN ECONÓMICA", ""),
        ("Ingresos Anuales Aproximados (COP) *", ""),
        ("¿Quién financia el viaje? *", ""),
        ("", ""),
        ("REFERENCIAS LABORALES", ""),
        ("Nombre del Supervisor/Jefe", ""),
        ("Teléfono del Supervisor", ""),
        ("Correo del Supervisor", ""),
    ]
    
    for label, default in campos_laboral:
        if label and not label.startswith("INFORMACIÓN") and not label.startswith("REFERENCIAS"):
            ws_laboral[f'A{row}'] = label
            ws_laboral[f'A{row}'].font = Font(bold=True)
            ws_laboral[f'B{row}'] = default
            ws_laboral[f'B{row}'].fill = PatternFill(start_color="FFFFFFE0", end_color="FFFFFFE0", fill_type="solid")
        elif label.startswith(("INFORMACIÓN", "REFERENCIAS")):
            ws_laboral[f'A{row}'] = label
            ws_laboral.merge_cells(f'A{row}:B{row}')
            ws_laboral[f'A{row}'].font = section_font
            ws_laboral[f'A{row}'].fill = section_fill
        row += 1
    
    # ======================== HOJA 7: FAMILIA ========================
    ws_familia = wb.create_sheet("Familia", 6)
    ws_familia.column_dimensions['A'].width = 35
    ws_familia.column_dimensions['B'].width = 40
    
    row = 1
    ws_familia[f'A{row}'] = "SECCIÓN 6: INFORMACIÓN FAMILIAR"
    ws_familia.merge_cells(f'A{row}:B{row}')
    ws_familia[f'A{row}'].font = header_font
    ws_familia[f'A{row}'].fill = header_fill
    ws_familia.row_dimensions[row].height = 20
    row += 2
    
    campos_familia = [
        ("Nombre Completo del Padre", ""),
        ("Fecha de Nacimiento del Padre", ""),
        ("¿Aún vive? (Sí/No)", ""),
        ("Nacionalidad del Padre", ""),
        ("", ""),
        ("Nombre Completo de la Madre", ""),
        ("Fecha de Nacimiento de la Madre", ""),
        ("¿Aún vive? (Sí/No)", ""),
        ("Nacionalidad de la Madre", ""),
        ("", ""),
        ("HERMANOS/AS", ""),
        ("¿Tienes hermanos? Cantidad:", ""),
        ("¿Alguno vive en EE.UU.? Detalles", ""),
        ("", ""),
        ("ESTADO CIVIL", ""),
        ("Estado Civil *", ""),
        ("¿Tienes Hijos? Cantidad:", ""),
        ("Edades de los Hijos", ""),
    ]
    
    for label, default in campos_familia:
        if label and not label.startswith(("HERMANOS", "ESTADO")):
            ws_familia[f'A{row}'] = label
            ws_familia[f'A{row}'].font = Font(bold=True)
            ws_familia[f'B{row}'] = default
            ws_familia[f'B{row}'].fill = PatternFill(start_color="FFFFFFE0", end_color="FFFFFFE0", fill_type="solid")
        elif label.startswith(("HERMANOS", "ESTADO")):
            ws_familia[f'A{row}'] = label
            ws_familia.merge_cells(f'A{row}:B{row}')
            ws_familia[f'A{row}'].font = section_font
            ws_familia[f'A{row}'].fill = section_fill
        row += 1
    
    # ======================== HOJA 8: DOCUMENTOS Y DECLARACIONES ========================
    ws_documentos = wb.create_sheet("Documentos", 7)
    ws_documentos.column_dimensions['A'].width = 60
    
    row = 1
    ws_documentos[f'A{row}'] = "SECCIÓN 7: DOCUMENTOS Y DECLARACIONES"
    ws_documentos[f'A{row}'].font = header_font
    ws_documentos[f'A{row}'].fill = header_fill
    ws_documentos.row_dimensions[row].height = 20
    row += 2
    
    ws_documentos[f'A{row}'] = "LISTA DE DOCUMENTOS REQUERIDOS"
    ws_documentos[f'A{row}'].font = section_font
    ws_documentos[f'A{row}'].fill = section_fill
    row += 1
    
    documentos = [
        "☐ Pasaporte válido (original, vigencia mínima 6 meses)",
        "☐ Cédula de ciudadanía original",
        "☐ Foto tamaño 5x5 cm (fondo blanco, cara frontal)",
        "☐ Confirmación de pago de la tarifa DS-160 ($160 USD)",
        "☐ Comprobante de medios económicos (últimos 3 meses):",
        "   - Extractos bancarios",
        "   - Cartas de empleador",
        "   - Comprobantes de ingresos",
        "☐ Comprobante de vivienda en Colombia (servicios públicos)",
        "☐ Si es visa B1: Carta del empleador en EE.UU.",
        "☐ Si es visa B2: Itinerario de viaje",
        "☐ Si es visa de estudiante: Carta de aceptación de la institución",
        "☐ Comprobante de fondos disponibles",
        "☐ Cualquier documento que demuestre vínculos con Colombia",
    ]
    
    for doc in documentos:
        ws_documentos[f'A{row}'] = doc
        row += 1
    
    row += 1
    ws_documentos[f'A{row}'] = "DECLARACIONES LEGALES"
    ws_documentos[f'A{row}'].font = section_font
    ws_documentos[f'A{row}'].fill = section_fill
    row += 1
    
    declaraciones = [
        "☐ DECLARO que toda la información es VERDADERA y COMPLETA",
        "☐ ENTIENDO que proporcionar información falsa es FRAUDE",
        "☐ ACEPTO que el fraude puede resultar en:",
        "   - Denegación permanente de visa",
        "   - Prohibición de entrada a EE.UU.",
        "   - Antecedentes penales",
        "   - Procesos legales internacionales",
        "☐ HE LEÍDO y ACEPTO todos los términos y condiciones",
        "☐ ENTIENDO que la visa puede ser denegada por cualquier motivo",
    ]
    
    for decl in declaraciones:
        ws_documentos[f'A{row}'] = decl
        row += 1
    
    # ======================== HOJA 9: RESUMEN Y CHECKLIST ========================
    ws_resumen = wb.create_sheet("Checklist Final", 8)
    ws_resumen.column_dimensions['A'].width = 60
    ws_resumen.column_dimensions['B'].width = 15
    
    row = 1
    ws_resumen[f'A{row}'] = "CHECKLIST FINAL - ANTES DE LA ENTREVISTA"
    ws_resumen.merge_cells(f'A{row}:B{row}')
    ws_resumen[f'A{row}'].font = header_font
    ws_resumen[f'A{row}'].fill = header_fill
    ws_resumen.row_dimensions[row].height = 20
    row += 2
    
    checklist_items = [
        ("He completado el formulario DS-160 oficial", "☐"),
        ("He pago la tarifa de solicitud", "☐"),
        ("He programado cita en la embajada", "☐"),
        ("He revisado toda la información para precisión", "☐"),
        ("He recopilado todos los documentos requeridos", "☐"),
        ("He llevado copia de todo al menos en 2 ocasiones", "☐"),
        ("He verificado vigencia de mi pasaporte (6+ meses)", "☐"),
        ("He impreso confirmación de pago (DS-160)", "☐"),
        ("He verificado horario y ubicación de la cita", "☐"),
        ("He preparado respuestas a posibles preguntas", "☐"),
        ("Voy a llegar temprano (15-30 min antes)", "☐"),
        ("He desactivado alarmas/notificaciones del celular", "☐"),
        ("Voy con documento de identidad en mano", "☐"),
        ("He fotografiado mis documentos (por si acaso)", "☐"),
    ]
    
    for item, checkbox in checklist_items:
        ws_resumen[f'A{row}'] = item
        ws_resumen[f'B{row}'] = checkbox
        ws_resumen[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1
    
    row += 2
    ws_resumen[f'A{row}'] = "NOTAS IMPORTANTES PARA LA ENTREVISTA"
    ws_resumen[f'A{row}'].font = section_font
    ws_resumen[f'A{row}'].fill = section_fill
    ws_resumen.merge_cells(f'A{row}:B{row}')
    row += 1
    
    notas = [
        "• Sé honesto y específico en todas tus respuestas",
        "• No inventes historias ni exageres",
        "• Mantén respuestas cortas y directas",
        "• Si no entiendes una pregunta, pide que la repita",
        "• Demuestra vínculos fuertes con Colombia",
        "• Prepara documentos en orden cronológico",
        "• Lleva dinero en efectivo (puede rechazarse tarjeta)",
        "• Sé puntual - llega 15-30 minutos antes",
        "• Viste de forma profesional y conservadora",
        "• No lleves electrónica (excepto lo permitido)",
    ]
    
    for nota in notas:
        ws_resumen[f'A{row}'] = nota
        ws_resumen.merge_cells(f'A{row}:B{row}')
        row += 1
    
    # Guardar
    archivo = f"Formulario_Visa_DS160_Colombianos_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
    wb.save(archivo)
    print(f"✅ Archivo Excel creado exitosamente: {archivo}")
    return archivo

if __name__ == "__main__":
    crear_formulario_visa()
